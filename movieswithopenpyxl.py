from openpyxl import load_workbook 


workbook = load_workbook(filename='file path')

# workbook.sheetnames


#workbook.sheetnames # data sheet

sheet=workbook.active


movies = {}
i=1  # id 
for row in sheet.iter_rows(min_row=2,
                           max_row=8,
                           min_col=3,
                           max_col=4,
                           values_only=True):


      movie_id=i
      movie={
         "movie name" : row[0],   
         "year":row[1],
         }
      i += 1
     
      movies[movie_id] = movie 

print(movies)
